import streamlit as st 
import numpy as np 
import os
import pandas as pd
import openpyxl
from copy import copy
import base64
import shutil
import re
import sys
from datetime import date
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill,Alignment,Side,Border

#危险源清单
class analyze_dangerlist:
    def __init__(self,database,type,name,date):
        self.dangerlist_path = os.path.abspath(r'templet/危险源清单模板.xlsx')
        self.dangerlist_save_path=os.path.abspath(r'result/危险源清单.xlsx')
        self.database = database
        self.type=type
        self.name=name
        self.date=date
    def change_database(self):
        if self.type=='国内':
            self.changed_database=self.database[self.database['国内航线']==1]
        elif self.type=='国际一般':
            self.changed_database=self.database[self.database['国际一般航线']==1]
        elif self.type=='国际特殊':
            self.changed_database=self.database[self.database['国际特殊航线']==1]
    def cell_update(self,addr,worksheet,values):
        cell_address = addr
        cell = worksheet[cell_address]
        # 从 addr 开始将每个值复制到相应的单元格中
        for i, value in enumerate(values):
            cell_to_update = worksheet.cell(row=cell.row+i, column=cell.column)
            cell_to_update.value=value
    def copy_row(self,worksheet,row_index,times):
        # 选择要复制的行
        row = worksheet[row_index]
        for n in range(1,times+1):
            for cell in row:
                new_cell=worksheet.cell(row=cell.row + n, column=cell.column)
                new_cell.value = cell.value
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.font = copy(cell.font)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)
    
    def new_dangerlist(self):
        self.change_database()
        data=self.changed_database
        [
            [
        '三级危险源','三级危险源潜在后果','三级危险源风险等级',
        '新原因1','新原因2','新原因3','新原因4','新原因5','新原因6','新原因7','新原因8','新原因9',
        '预防措施1','预防措施2','预防措施3','预防措施4','预防措施5','预防措施6','预防措施7','预防措施8',
        '应急措施1','应急措施2','应急措施3'
        ]
        ]
        # 将新原因1-9列合并为新原因列
        # 如果只有一个新原因，直接使用该原因作为新原因列
        if len(data.columns[data.columns.str.contains('新原因')]) == 1:
            data['新原因'] = data['新原因1']
        else:
            # 否则，将新原因1-9列合并为新原因列
            new_reasons = data[['新原因1', '新原因2', '新原因3', '新原因4', '新原因5', '新原因6', '新原因7', '新原因8', '新原因9']]
            new_reasons_str = new_reasons.apply(lambda x: ' '.join([f'{i+1}、{v}' for i, v in enumerate(x) if pd.notna(v)]), axis=1)
        # 将预防措施1-8列和应急措施1-3列合并为措施列
        precautions = data[['预防措施1', '预防措施2', '预防措施3', '预防措施4', '预防措施5', '预防措施6', '预防措施7', '预防措施8']]
        emergency_measures = data[['应急措施1', '应急措施2', '应急措施3']]
        measures = precautions.join(emergency_measures)
        measures_str = measures.apply(lambda x: ' '.join([f'{i+1}、{v}' for i, v in enumerate(x) if pd.notna(v)]), axis=1)

        # 将新原因1-9列和预防措施1-8列和应急措施1-3列合并为新 DataFrame
        new_database = pd.DataFrame({
            '三级危险源': data['三级危险源'],
            '三级危险源潜在后果': data['三级危险源潜在后果'],
            '三级危险源风险等级': data['三级危险源风险等级'],
            '新原因': new_reasons_str,
            '措施': measures_str
        })
        #加载 Excel 文件
        workbook=openpyxl.load_workbook(self.dangerlist_path)
        ws = workbook.active
        # 复制数据
        row_index=5
        self.copy_row(ws,row_index,new_database.shape[0]-1)
        #对照数据库修改
        self.cell_update('E5',ws,new_database['三级危险源'])
        self.cell_update('G5',ws,new_database['三级危险源潜在后果'])
        self.cell_update('J5',ws,new_database['三级危险源风险等级'])
        self.cell_update('L5',ws,new_database['新原因'])
        self.cell_update('M5',ws,new_database['措施'])

        # 获取单元格内容并替换姓名和分析日期
        cell_value = ws["A2"].value
        name_pattern = r"分析人：\s*(\S{2,4})\s"
        name_match = re.search(name_pattern, cell_value)
        if name_match:
            name = name_match.group(1)
            # 替换为您要设置的姓名
            new_name_str = self.name
            cell_value = cell_value.replace(name,new_name_str )
        date_pattern = r"分析日期：\s*(\d{4}年\d{1,2}月\d{1,2}日)"
        date_match = re.search(date_pattern, cell_value)
        if date_match:
            date_str = date_match.group(1)
            # 替换为您要设置的日期
            new_date_str = self.date
            cell_value = cell_value.replace(date_str, new_date_str)
        
        # 将替换后的值写回单元格
        ws["A2"].value = cell_value
        #修改审批日期
        cell_value = ws["A3"].value
        date_pattern = r"审批日期：\s*(\d{4}年\d{1,2}月\d{1,2}日)"
        date_match = re.search(date_pattern, cell_value)
        if date_match:
            date_str = date_match.group(1)
            # 替换为您要设置的日期
            new_date_str = self.date
            cell_value = cell_value.replace(date_str, new_date_str)
        # 将替换后的值写回单元格
        ws["A3"].value = cell_value
        #保存
        workbook.save(self.dangerlist_save_path)
    def main(self):
        self.new_dangerlist()

#风险评价报告表
class analyze_report:
    def __init__(self,database,type,name,date,title):
        self.report_path = os.path.abspath(r'templet/风险评价报告表模版.xlsx')
        self.report_save_path=os.path.abspath(r'result/风险评价报告表.xlsx')
        self.database = database
        self.type=type
        self.name=name
        self.date=date
        self.title=title
        self.workbook=load_workbook(self.report_path)
        self.ws=self.workbook.active
    def change_database(self):
        if self.type=='国内':
            self.changed_database=self.database[self.database['国内航线']==1]
        elif self.type=='国际一般':
            self.changed_database=self.database[self.database['国际一般航线']==1]
        elif self.type=='国际特殊':
            self.changed_database=self.database[self.database['国际特殊航线']==1]
        new_reasons = self.changed_database[['新原因1', '新原因2', '新原因3', '新原因4', '新原因5', '新原因6', '新原因7', '新原因8', '新原因9']]
        new_reasons_str = new_reasons.apply(lambda x: ' '.join([f'{i+1}、{v}' for i, v in enumerate(x) if pd.notna(v)]), axis=1)
        # 将预防措施1-8列和应急措施1-3列合并为措施列
        precautions = self.changed_database[['预防措施1', '预防措施2', '预防措施3', '预防措施4', '预防措施5', '预防措施6', '预防措施7', '预防措施8']]
        emergency_measures = self.changed_database[['应急措施1', '应急措施2', '应急措施3']]
        measures = precautions.join(emergency_measures)
        measures_str = measures.apply(lambda x: ' '.join([f'{i+1}、{v}' for i, v in enumerate(x) if pd.notna(v)]), axis=1)
        self.new_database = pd.DataFrame({
                    '三级危险源':self.changed_database['三级危险源'],
                    '新原因': new_reasons_str,
                    '控制措施': measures_str,
                    '严重性分析': self.changed_database['三级危险源严重性等级'],
                    '严重性分值': self.changed_database['三级危险源严重性分值'],
                    '可能性分析': self.changed_database['三级危险源可能性等级'],
                    '可能性分值': self.changed_database['三级危险源可能性分值'],
                    '风险等级': self.changed_database['三级危险源风险等级'],
                    '风险值': self.changed_database['三级危险源风险分值'],
                })
    #part0:表头、
    def part_title(self):
        part_rows=[]
        title1_row=[cell.value for cell in self.ws[1]]
        title2_row=[cell.value for cell in self.ws[2]]
        title3_row=[cell.value for cell in self.ws[3]]
        title4_row=[cell.value for cell in self.ws[4]]
        title5_row=[cell.value for cell in self.ws[5]]
        new_title1_row=title1_row.copy()
        new_title2_row=title2_row.copy()
        new_title2_row[1]=self.title
        new_title3_row=title3_row.copy()
        new_title3_row[5]=self.date
        new_title4_row=title4_row.copy()
        new_title5_row=title5_row.copy()
        new_title5_row[3]=self.name
        part_rows.append([new_title1_row,new_title2_row,new_title3_row,new_title4_row,new_title5_row])
        return part_rows
    #part1:危险源识别
    def part_1(self):
        part_rows=[]
        page_row=[cell.value for cell in self.ws[6]]
        part_row=[cell.value for cell in self.ws[7]]
        title_row=[cell.value for cell in self.ws[8]]
        modify_row=[cell.value for cell in self.ws[9]]
        for x in range(self.new_database.shape[0]):
            new_page_row=page_row.copy()
            new_part_row=part_row.copy()
            new_title_row=title_row.copy()
            new_modify_row=modify_row.copy()

            page=new_page_row[0]
            page=re.sub(r'共\s+页',f'共{str(self.new_database.shape[0])}页',page)
            page=re.sub(r'第\s+页',f'第{str(x+1)}页',page)
            new_page_row[0]=page
            text=self.new_database['新原因'].tolist()[x]
            new_modify_row[1]=f'背景描述：航班运行中发生{text}未及时处置。\n危险源识别：{text}'
            part_rows.append([new_page_row,new_part_row,new_title_row,new_modify_row])
        return part_rows
    #part2:风险分析和评价
    def part_2(self):
        #读取
        part_rows=[]
        part_row=[cell.value for cell in self.ws[10]]
        reason_row=[cell.value for cell in self.ws[11]]
        modify_import_row=[cell.value for cell in self.ws[12]]
        modify_probability_row=[cell.value for cell in self.ws[13]]
        modify_risk_row=[cell.value for cell in self.ws[14]]
        #原因分析

        for x in range(self.new_database.shape[0]):
            #创建copy
            new_part_row=part_row.copy()
            new_reason_row=reason_row.copy()
            new_modify_import_row=modify_import_row.copy()
            new_modify_probability_row=modify_probability_row.copy()
            new_modify_risk_row=modify_risk_row.copy()

            new_reason_row[1]=self.new_database['新原因'].tolist()[x]
            new_modify_import_row[1]=self.new_database['严重性分析'].tolist()[x]
            new_modify_import_row[4]=self.new_database['严重性分值'].tolist()[x]
            new_modify_probability_row[1]=self.new_database['可能性分析'].tolist()[x]
            new_modify_probability_row[4]=self.new_database['可能性分值'].tolist()[x]
            new_modify_risk_row[1]=self.new_database['风险等级'].tolist()[x]
            new_modify_risk_row[4]=self.new_database['风险值'].tolist()[x]
            part_rows.append([new_part_row,new_reason_row, new_modify_import_row, new_modify_probability_row, new_modify_risk_row])
        return part_rows
    #part3:风险控制
    def part_3(self):
        #读取
        part_rows=[]
        part_row=[cell.value for cell in self.ws[15]]
        respon_unit_row=[cell.value for cell in self.ws[16]]
        account_unit_row=[cell.value for cell in self.ws[17]]
        target_row=[cell.value for cell in self.ws[18]]
        modify_measures_row=[cell.value for cell in self.ws[19]]
        for x in range(self.new_database.shape[0]):
            #创建copy
            new_part_row=part_row.copy()
            new_respon_unit_row=respon_unit_row.copy()
            new_account_unit_row=account_unit_row.copy()
            new_target_row=target_row.copy()
            new_modify_measures_row=modify_measures_row.copy()

            new_modify_measures_row[1]=self.new_database['控制措施'].tolist()[x]
            part_rows.append([new_part_row,new_respon_unit_row, new_account_unit_row,new_target_row, new_modify_measures_row])
        return part_rows
    
    def new_report(self):
        new_workbook=openpyxl.Workbook()
        worksheet = new_workbook.active  # 获取默认的工作表
        worksheet.title = 'Sheet1'   # 重命名工作表
        
        part0=self.part_title()
        for mudule in part0:
            for line in mudule:
                worksheet.append(line)
        part1=self.part_1()
        part2=self.part_2()
        part3=self.part_3()
        results=[]
        for n in range(self.new_database.shape[0]):
            results.append([part1[n],part2[n],part3[n]])
        
        for mudule in results:
            for item in mudule:
                for line in item:
                    new_result_line=[cell if cell is not None else '' for cell in line]
                    
                    worksheet.append(new_result_line)

        new_workbook.save(self.report_save_path)
    #修改格式
    def change_styles(self):
        workbook=load_workbook(self.report_save_path)
        ws=workbook.active
        cols=('A','B','C','D','E','F')
        font = Font(
                name="宋体",
                size=11,
                bold=False,
                italic=False,
                strike=None,
                underline=None,
            )
        side1 = Side(style="thin",color="000000")
        #设置字体,对齐方式，边框
        for col in cols:
            # 从A列至F列
            for i in range(1,ws.max_row+1):
                ws[f'{col}{i}'].font = font
                ws[f'{col}{i}'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                #设置边框
                ws[f'{col}{i}'].border = Border(left=side1,right=side1,top=side1,bottom=side1)
        #设置行高
        for row in ws.rows:
            for cell in row:
                ws.row_dimensions[cell.row].height=14.4
        #合并单元格
        for n in range(6,ws.max_row+1,14):
            ws.merge_cells(f'A{n}:B{n}')
            ws.merge_cells(f'D{n}:F{n}')
            ws.merge_cells(f'A{n+1}:F{n+1}')
            ws.merge_cells(f'B{n+2}:F{n+2}')
            ws.merge_cells(f'B{n+3}:F{n+3}')

            ws.merge_cells(f'A{n+4}:F{n+4}')
            ws.merge_cells(f'B{n+5}:F{n+5}')
            ws.merge_cells(f'B{n+6}:C{n+6}')
            ws.merge_cells(f'E{n+6}:F{n+6}')
            ws.merge_cells(f'B{n+7}:C{n+7}')
            ws.merge_cells(f'E{n+7}:F{n+7}')
            ws.merge_cells(f'B{n+8}:C{n+8}')
            ws.merge_cells(f'E{n+8}:F{n+8}')

            ws.merge_cells(f'A{n+9}:F{n+9}')
            ws.merge_cells(f'B{n+10}:F{n+10}')
            ws.merge_cells(f'B{n+11}:F{n+11}')
            ws.merge_cells(f'B{n+12}:F{n+12}')
        #标题部分格式和列宽处理
        ws['B2'].font=Font(name="宋体",size=11,bold=True,italic=False,strike=None,underline=None)
        ws.merge_cells('A1:F1')
        ws.merge_cells('A2:A4')
        ws.merge_cells('B2:D4')
        # 调整列宽
        ws.column_dimensions['A'].width = 12.4
        ws.column_dimensions['B'].width = 17
        ws.column_dimensions['C'].width = 6
        ws.column_dimensions['D'].width = 12.4
        ws.column_dimensions['E'].width = 12.4
        ws.column_dimensions['F'].width = 17
        ws.delete_cols(7,10)
        workbook.save(self.report_save_path)
    def main(self):
        self.change_database()
        self.new_report()
        self.change_styles()

#系统与工作分析记录表
class analyze_sysrecord:
    def __init__(self,database,type,name,date):
        self.sysrecord_path = os.path.abspath(r'templet/系统与工作分析记录表模板.xlsx')
        self.sysrecord_save_path=os.path.abspath(r'result/系统与工作分析记录表.xlsx')
        self.database = database
        self.type=type
        self.name=name
        self.date=date
    def change_database(self):
        if self.type=='国内':
            self.changed_database=self.database[self.database['国内航线']==1]
        elif self.type=='国际一般':
            self.changed_database=self.database[self.database['国际一般航线']==1]
        elif self.type=='国际特殊':
            self.changed_database=self.database[self.database['国际特殊航线']==1]
    def new_sysrecord(self):
        workbook=openpyxl.load_workbook(self.sysrecord_path)
        ws=workbook.active
        # 获取单元格内容并替换姓名和分析日期
        cell_value = ws["A2"].value
        name_pattern = r"分析人：\s*(\S{2,4})\s"
        name_match = re.search(name_pattern, cell_value)
        if name_match:
            name = name_match.group(1)
            # 替换为您要设置的姓名
            new_name_str = self.name
            cell_value = cell_value.replace(name,new_name_str )
        date_pattern = r"分析日期：\s*(\d{4}年\d{1,2}月\d{1,2}日)"
        date_match = re.search(date_pattern, cell_value)
        if date_match:
            date_str = date_match.group(1)
            # 替换为您要设置的日期
            new_date_str = self.date
            cell_value = cell_value.replace(date_str, new_date_str)
        # 将替换后的值写回单元格
        ws["A2"].value = cell_value
        
        #修改审批日期
        cell_value = ws["A3"].value
        date_pattern = r"审批日期：\s*(\d{4}年\d{1,2}月\d{1,2}日)"
        date_match = re.search(date_pattern, cell_value)
        if date_match:
            date_str = date_match.group(1)
            # 替换为您要设置的日期
            new_date_str = self.date
            cell_value = cell_value.replace(date_str, new_date_str)
        # 将替换后的值写回单元格
        ws["A3"].value = cell_value
        #保存
        workbook.save(self.sysrecord_save_path)
    def main(self):
        self.new_sysrecord()

def form_callback():
    if 'database' not in st.session_state:
        st.warning('数据库未导入')
    elif 'name' not in st.session_state:
        st.warning('请输入姓名')
    elif 'title' not in st.session_state:
        st.warning('请输入标题')
    elif 'datestr' not in st.session_state:
        st.warning('请输入日期')
    else:
        st.write('Data Saved')
        st.session_state.datasavecode=True
def download_button(file_path, button_text):
    with open(os.path.abspath(file_path), 'rb') as f:
        bytes = f.read()
        b64 = base64.b64encode(bytes).decode()

    # 创建一个名为 "Download File" 的下载链接
    href = f'<a href="data:application/octet-stream;base64,{b64}" download="{os.path.basename(file_path)}">{button_text}</a>'

    # 在 Streamlit 应用程序中使用按钮链接
    st.markdown(f'<div class="button-container">{href}</div>', unsafe_allow_html=True)

    # 添加 CSS 样式以将链接样式化为按钮
    st.markdown("""
        <style>
        .button-container {
            display: inline-block;
            margin-top: 1em;
        }
        .button-container a {
            background-color: #0072C6;
            border: none;
            color: white;
            padding: 0.5em 1em;
            text-align: center;
            text-decoration: none;
            display: inline-block;
            font-size: 16px;
            font-weight: bold;
            border-radius: 4px;
            cursor: pointer;
        }
        .button-container a:hover {
            background-color: #005AA3;
        }
        </style>
    """, unsafe_allow_html=True)
if __name__ == "__main__":
    st.set_page_config(page_title="analyze", page_icon="✈", layout="wide")

    with st.form(key='my_form'):
        flight_type = st.selectbox('Select Flight Type', ['国内', '国际一般', '国际特殊'], key='flight_type')
        # 创建输入日期的对话框
        input_date=st.date_input('请输入日期',date.today(), key='input_date')
        st.session_state.datestr=input_date.strftime("%Y年%m月%d日")
        # 创建输入姓名的对话框
        name = st.text_input("请输入您的姓名：", key='name')
        # 创建输入标题的对话框
        title = st.text_input("请输入标题：",'xx机型运行xx至xx机场往返航线风险评价', key='title')
        #checkbox_input = st.checkbox('Yes or No', key='my_checkbox')
        submit_button = st.form_submit_button(label='Submit', on_click=form_callback)

    #页面设置
    st.title(st.session_state.flight_type+'航线新开航分析')
    #危险源清单
    with st.container():
        st.write('-------------------------')
        st.subheader('危险源清单')
        left_column, right_column = st.columns(2)
        with left_column:
            if st.button('生成危险源清单'):
                with st.spinner('正在处理数据，请稍等...'):
                    if 'datasavecode' not in st.session_state:
                        st.warning('初始数据未准备正确,请上传数据文件')
                    else:
                        #实例化方法
                        dangerlist=analyze_dangerlist(st.session_state.database,st.session_state.flight_type,st.session_state.name,st.session_state.datestr)
                        dangerlist.main()
                        st.write('complete')
                        with right_column:
                            download_button(os.path.abspath(r'result/危险源清单.xlsx'), '下载危险源清单')
                        
    #风险评价报告表
    with st.container():
        st.write('-------------------------')
        st.subheader('风险评价报告表')
        left_column, right_column = st.columns(2)
        with left_column:
            if st.button('风险评价报告表'):
                with st.spinner('正在处理数据，请稍等...'):
                    if 'datasavecode' not in st.session_state:
                        st.warning('初始数据未准备正确,请上传数据文件')
                    else:
                        #实例化方法
                        report=analyze_report(st.session_state.database,st.session_state.flight_type,st.session_state.name,st.session_state.datestr,st.session_state.title)
                        report.main()
                        st.write('complete')
                        with right_column:
                            download_button(os.path.abspath(r'result/风险评价报告表.xlsx'), '下载风险评价报告表')
    #系统与工作分析记录表
    with st.container():
        st.write('-------------------------')
        st.subheader('系统与工作分析记录表')
        left_column, right_column = st.columns(2)
        with left_column:
            if st.button('系统与工作分析记录表'):
                with st.spinner('正在处理数据，请稍等...'):
                    if 'datasavecode' not in st.session_state:
                        st.warning('初始数据未准备正确,请上传数据文件') 
                    else:
                        #实例化方法
                        sysrecord=analyze_sysrecord(st.session_state.database,st.session_state.flight_type,st.session_state.name,st.session_state.datestr)
                        sysrecord.main()
                        st.write('complete')
                        with right_column:
                            download_button(os.path.abspath(r'result/系统与工作分析记录表.xlsx'), '系统与工作分析记录表')
                    
